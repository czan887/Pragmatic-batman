"""
File import API routes for loading data from files
"""
from fastapi import APIRouter, UploadFile, File, HTTPException, status
from pydantic import BaseModel
from typing import Optional
import io
import csv

router = APIRouter()


class ImportResult(BaseModel):
    """File import result"""
    status: str
    items: list[str]
    count: int
    filename: str
    format: str


def parse_txt_content(content: str) -> list[str]:
    """Parse text file content (one item per line)"""
    items = []
    for line in content.splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            # Remove @ prefix if present (for usernames)
            if line.startswith("@"):
                line = line[1:]
            items.append(line)
    return items


def parse_csv_content(content: str) -> list[str]:
    """Parse CSV file content (first column)"""
    items = []
    reader = csv.reader(io.StringIO(content))

    # Skip header if it looks like one
    first_row = next(reader, None)
    if first_row:
        # Check if first value looks like a header
        first_val = first_row[0].lower() if first_row else ""
        header_keywords = ["username", "user", "url", "link", "hashtag", "id", "name"]
        if not any(keyword in first_val for keyword in header_keywords):
            # Not a header, include it
            val = first_row[0].strip()
            if val.startswith("@"):
                val = val[1:]
            if val:
                items.append(val)

    for row in reader:
        if row:
            val = row[0].strip()
            if val.startswith("@"):
                val = val[1:]
            if val:
                items.append(val)

    return items


async def parse_xlsx_content(content: bytes) -> list[str]:
    """Parse Excel file content (first column)"""
    try:
        import openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active

        items = []
        first_row = True

        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            if row[0] is not None:
                val = str(row[0]).strip()

                # Skip header row
                if first_row:
                    first_row = False
                    header_keywords = ["username", "user", "url", "link", "hashtag", "id", "name"]
                    if any(keyword in val.lower() for keyword in header_keywords):
                        continue

                if val.startswith("@"):
                    val = val[1:]
                if val:
                    items.append(val)

        wb.close()
        return items

    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel support requires openpyxl. Please install it: pip install openpyxl"
        )


@router.post("/", response_model=ImportResult)
async def import_file(file: UploadFile = File(...)):
    """
    Import data from file

    Supports:
    - .txt files (one item per line)
    - .csv files (uses first column)
    - .xlsx/.xls files (uses first column)

    Returns list of extracted items (usernames, URLs, hashtags, etc.)
    """
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No filename provided"
        )

    # Determine file format
    filename = file.filename.lower()
    if filename.endswith(".txt"):
        file_format = "txt"
    elif filename.endswith(".csv"):
        file_format = "csv"
    elif filename.endswith((".xlsx", ".xls")):
        file_format = "xlsx"
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file format. Supported: .txt, .csv, .xlsx, .xls"
        )

    # Read file content
    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to read file: {e}"
        )

    # Parse based on format
    try:
        if file_format == "txt":
            items = parse_txt_content(content.decode("utf-8", errors="ignore"))
        elif file_format == "csv":
            items = parse_csv_content(content.decode("utf-8", errors="ignore"))
        elif file_format == "xlsx":
            items = await parse_xlsx_content(content)
        else:
            items = []
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {e}"
        )

    if not items:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No valid items found in file"
        )

    # Remove duplicates while preserving order
    seen = set()
    unique_items = []
    for item in items:
        if item not in seen:
            seen.add(item)
            unique_items.append(item)

    return ImportResult(
        status="success",
        items=unique_items,
        count=len(unique_items),
        filename=file.filename,
        format=file_format
    )


@router.post("/validate")
async def validate_file(file: UploadFile = File(...)):
    """
    Validate file format without fully parsing

    Quick check to verify file can be parsed.
    """
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No filename provided"
        )

    filename = file.filename.lower()
    supported = filename.endswith((".txt", ".csv", ".xlsx", ".xls"))

    if not supported:
        return {
            "valid": False,
            "filename": file.filename,
            "error": "Unsupported file format"
        }

    # Try to read first few bytes
    try:
        content = await file.read(1024)
        await file.seek(0)

        return {
            "valid": True,
            "filename": file.filename,
            "size_bytes": len(content),
            "format": filename.split(".")[-1]
        }
    except Exception as e:
        return {
            "valid": False,
            "filename": file.filename,
            "error": str(e)
        }
